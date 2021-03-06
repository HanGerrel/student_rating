from datetime import datetime
from uuid import uuid4

import django.contrib.auth as da
from django.conf import settings
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.core.paginator import Paginator
from django.core.serializers import serialize
from django.http import JsonResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook

import main.forms as forms
import main.models as model


def winter(year):
    start = year + "-08-02"
    end = str(int(year) + 1) + "-02-28"
    return [start, end]


def summer(year):
    start = year + "-03-01"
    end = year + "-08-01"
    return [start, end]


def get_rating(faculty: int, session, year):
    date = []
    if session == "summer":
        date = summer(year)
    if session == "winter":
        date = winter(year)
    rating = model.Rating.objects.filter(date__range=date, faculty=faculty).values('id', 'full_name', 'group',
                                                                                   'session', 'extra', 'total')
    return rating.order_by('-total')


def index(request):
    form = forms.FilterForm(request.GET)
    if form.is_valid():
        faculty = form.cleaned_data.get('faculty')
        session = form.cleaned_data.get('session')
        year = form.cleaned_data.get('year_picker')
        dysplayed = form.cleaned_data.get('dysplayed')
        rating = get_rating(faculty, session, year)
    else:
        form = forms.FilterForm(initial={'year_picker': str(datetime.now().year)})
        faculty = form['faculty'].initial
        session = form['session'].initial
        year = form['year_picker'].initial
        dysplayed = form['dysplayed'].initial
        rating = get_rating(faculty, session, year)

    paginator = Paginator(rating, dysplayed)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {"page_obj": page_obj, "form": form}
    return render(request, 'main/index.html', context)


def upload_certificate(request):
    content_type = ["image/jpeg", "image/jpg", "image/png"]
    file = request.FILES['upload_file']
    print(request.POST['student_id'])
    if file.content_type in content_type:
        student = model.Rating.objects.get(pk=request.POST['student_id'])
        table_entry = model.Certificate(uploaded_by_student=student, certificate_file=file)
        table_entry.save()
    return redirect('/')


def get_details(request):
    if request.is_ajax():
        student_id = request.GET['student']
        details = model.ExtraPoint.objects.filter(student_id=student_id)
        details = serialize('json', details)
        return JsonResponse(details, safe=False)


def login(request):
    context = {'error': ''}
    if request.method == "POST":
        username = request.POST['login']
        password = request.POST['password']
        user = da.authenticate(request, username=username, password=password)
        if user is not None:
            da.login(request, user)
            return redirect('/')
        else:
            context = {'error': 'invalid login'}
    return render(request, 'main/sign_in.html', context)


def sign_up(request):
    da.logout(request)
    invite_key_status = ""
    error = ""
    if request.method == 'POST':
        if request.POST['invite_key_status'] != 'OK':
            invite_key = request.POST['invite_key']
            if model.InviteKey.objects.filter(invite_key__exact=invite_key).count():
                invite_key_status = "OK"
            else:
                error = 'invalid key'
        else:
            full_name = request.POST['full_name'].split(' ')
            first_name = full_name[0]
            last_name = ''
            if len(full_name) > 1:
                last_name = full_name[1]
            if len(first_name) == 0 or len(last_name) == 0:
                error = 'enter your full name'
                context = {"invite_key": "OK", "error": error}
                return render(request, 'main/sign_up.html', context)

            faculty = request.POST.get('faculty')
            if faculty is None:
                error = 'enter your faculty'
                context = {"invite_key": "OK", "error": error}
                return render(request, 'main/sign_up.html', context)

            email = request.POST['email']
            login = request.POST['login']
            password = request.POST['password']

            user = User.objects.create_user(login, email, password, first_name=first_name, last_name=last_name)
            user.save()
            return redirect('/')
    context = {"invite_key": invite_key_status, "error": error}
    return render(request, 'main/sign_up.html', context)


def invite_key_gen(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            invite_key = ""
            if request.method == "POST":
                invite_key = uuid4()
                a_record = model.InviteKey(invite_key=invite_key)
                a_record.save()
            context = {"invite_key": invite_key}
            return render(request, 'admin/invite_key_gen.html', context)
    return redirect('/')


def add_rating(request):
    username = request.user.username
    content_type = ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel"]
    if request.method == "POST":
        if 'upload_rating' in request.POST:
            file = request.FILES['upload_file']
            if file.content_type in content_type:
                table_entry = model.ExelFile(uploaded_by_user=username, excel_file=file)
                table_entry.save()
                read_workbook(file.name, 'new')
            return redirect('/add-rating')
        if 'add_rating' in request.POST:
            form = forms.AddRatingForm(request.POST)
            if form.is_valid():
                rating = form.save(commit=False)
                rating.total = form.cleaned_data.get('session') + form.cleaned_data.get('extra')
                rating.save()
                return redirect('/')
            else:
                context = {'username': username, "form": form}
                return render(request, 'main/add_rating.html', context)
    else:
        form = forms.AddRatingForm()
        context = {'username': username, "form": form}
        return render(request, 'main/add_rating.html', context)


def read_workbook(file_name, action):
    file_dir = str(settings.MEDIA_ROOT) + "/media/excel/"
    workbook = load_workbook(file_dir + file_name)
    worksheet = workbook[workbook.sheetnames[0]]
    faculties = {
        "ФИТ": 1,
        "ФІТ": 1,
        "ЕкФ": 2,
        "ЕнФ": 3,
        "МФ": 4,
        "СГФ": 5,
        "ФМЗ": 6,
        "ФТТ": 7,
        "ФИМП": 8,
        "ФІМП": 8,
    }
    faculty = faculties.get(worksheet.cell(row=2, column=2).value)
    faculty = model.Faculty.objects.get(id=faculty)

    if action == 'new':
        for row in worksheet.iter_rows(min_row=2):
            temp = []
            for cells in row:
                temp.append(cells.value)
            table_entry = model.Rating()
            table_entry.full_name = temp[0]
            table_entry.faculty = faculty
            table_entry.group = temp[2]
            table_entry.session = temp[3]
            table_entry.extra = temp[4]
            table_entry.total = float(temp[3]) + float(temp[4])
            table_entry.save()
    if action == 'update':
        for row in worksheet.iter_rows(min_row=2):
            temp = []
            for cells in row:
                temp.append(cells.value)
            table_entry = model.Rating.objects.get(faculty=faculty, group=temp[2], full_name=temp[0])
            table_entry.session = temp[3]
            table_entry.extra = temp[4]
            table_entry.total = float(temp[3]) + float(temp[4])
            table_entry.save()


def check_certificate(request):
    if request.is_ajax():
        record_id = request.POST['record_id']
        if request.POST['action'] == 'add':
            certificate = model.Certificate.objects.get(pk=record_id)
            student_id = certificate.uploaded_by_student
            point = request.POST['added_points']
            description = request.POST['activity']
            certificate_file = certificate.certificate_file
            extra_points = model.ExtraPoint(student_id=student_id, point=point, description=description,
                                            certificate=certificate_file)
            student_rating = model.Rating.objects.get(pk=student_id.id)
            student_rating.extra += int(point)
            student_rating.total += int(point)
            student_rating.save()
            extra_points.save()
            certificate.delete()
        if request.POST['action'] == 'reject':
            model.Certificate.objects.get(pk=record_id).delete()

    certificates = model.Certificate.objects.all()
    context = {'certificates': certificates}
    return render(request, 'main/check-certificate.html', context)


def change_rating(request):
    faculties = model.Faculty.objects.all()
    student = model.Rating.objects.filter(faculty=1)
    if request.is_ajax():
        if 'faculty' in request.GET:
            student = model.Rating.objects.filter(faculty=int(request.GET['faculty']))
        if 'student' in request.GET:
            student = model.Rating.objects.filter(pk=int(request.GET['student']))
        student = serialize('json', student)
        return JsonResponse(student, safe=False)
    if request.method == "POST":
        student_id = int(request.POST['student_id'])
        session = float(request.POST['session'])
        extra_point = int(request.POST['added_points'])
        description = request.POST['activity']
        file = ''
        if 'upload_file' in request.FILES:
            file = request.FILES['upload_file']
        student_rating = model.Rating.objects.get(pk=student_id)
        extra_points = model.ExtraPoint(student_id=student_rating, point=extra_point, description=description,
                                        certificate=file)
        student_rating.session = session
        student_rating.extra += extra_point
        student_rating.total += extra_points
        student_rating.save()
        extra_points.save()

    context = {'faculties': faculties, 'students': student}
    return render(request, 'main/change_rating.html', context)


def change_from_file(request):
    username = request.user.username
    content_type = ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel"]
    if request.method == "POST":
        file = request.FILES['upload_file']
        if file.content_type in content_type:
            table_entry = model.ExelFile(uploaded_by_user=username, excel_file=file)
            table_entry.save()
            read_workbook(file.name, 'update')
    return redirect('home')


def profile(request):
    success = ''
    password_change_form = PasswordChangeForm(user=request.user)
    context = {}
    if request.is_ajax():
        if request.user.is_superuser:
            invite_key = uuid4()
            a_record = model.InviteKey(invite_key=invite_key)
            a_record.save()
            context = {"invite_key": invite_key}
        return JsonResponse(context)
    if request.method == 'POST':
        user = User.objects.get(id=request.user.id)
        if 'update_profile' in request.POST:
            user.username = request.POST['username']
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.save()
            success = 'profile updated successfully'
        if 'update_email' in request.POST:
            print(request.POST['email'])
            user.email = request.POST['email']
            user.save()
            success = 'email updated successfully'
        if 'update_password' in request.POST:
            form = PasswordChangeForm(user=request.user, data=request.POST or None)
            print('test')
            if form.is_valid():
                print('test2')
                form.save()
                update_session_auth_hash(request, form.user)
            success = 'password updated successfully'
        context = {"success": success}
        return render(request, 'main/profile.html', context)
    context = {"password_change_form": password_change_form, "success": success}
    return render(request, 'main/profile.html', context)
